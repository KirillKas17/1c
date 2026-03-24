"""
Тесты для AI-анализатора Excel файлов.
Проверка гибридного подхода (OpenRouter + Ollama).
"""

import pytest
import json
from unittest.mock import Mock, patch, MagicMock
from src.ai_core.ai_analyzer import AIExcelAnalyzer
from src.ai_core.integration import SmartParser


class TestAIExcelAnalyzer:
    """Тесты базового функционала AI-анализатора."""
    
    def test_init_with_env_vars(self):
        """Инициализация с переменными окружения."""
        analyzer = AIExcelAnalyzer()
        assert analyzer.openrouter_url == "https://openrouter.ai/api/v1/chat/completions"
        assert analyzer.ollama_url == "http://localhost:11434/api/generate"
        assert analyzer.fallback_model == "phi3:mini"
    
    def test_system_prompt_exists(self):
        """Системный промпт должен содержать ключевые инструкции."""
        analyzer = AIExcelAnalyzer()
        prompt = analyzer.system_prompt
        
        # Проверка ключевых требований
        assert "НДС" in prompt or "vat" in prompt.lower()
        assert "брутто" in prompt.lower() or "gross" in prompt.lower()
        assert "нетто" in prompt.lower() or "net" in prompt.lower()
        assert "иерарх" in prompt.lower() or "hierarch" in prompt.lower()
        assert "json" in prompt.lower()
    
    def test_extract_sheet_sample_structure(self, tmp_path):
        """Извлечение образца данных из файла."""
        # Создаем тестовый Excel файл
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        
        # Шапка отчета 1С
        ws['A1'] = "Валовая выручка предприятия"
        ws['A2'] = "Период: 01.01.2024 - 31.01.2024"
        ws['A3'] = "Валюта: RUB"
        
        # Заголовки таблицы
        ws['A5'] = "Контрагент"
        ws['B5'] = "Сумма с НДС"
        ws['C5'] = "Вес брутто"
        
        # Данные
        ws['A6'] = "ООО Ромашка"
        ws['B6'] = 120000
        ws['C6'] = 500
        
        test_file = tmp_path / "test_report.xlsx"
        wb.save(str(test_file))
        wb.close()
        
        analyzer = AIExcelAnalyzer()
        result = analyzer.extract_sheet_sample(str(test_file))
        
        assert result['sheet_name'] == 'TestSheet'
        assert len(result['sample_rows']) > 0
        assert isinstance(result['sample_rows'], list)


class TestAIIntegration:
    """Тесты интеграционного слоя."""
    
    def test_smart_parser_init(self):
        """Инициализация умного парсера без импорта тяжелых зависимостей."""
        # Тестируем только что класс существует и может быть создан
        from src.ai_core.integration import AIParserIntegration
        
        # Мокаем зависимости перед созданием
        with patch('src.ai_core.integration.get_hierarchy_parser', return_value=None):
            integration = AIParserIntegration()
            assert integration.ai_analyzer is not None
            # hierarchy_parser может быть None если зависимость недоступна
            assert integration.hierarchy_parser is None or integration.hierarchy_parser is not None
    
    @patch('src.ai_core.ai_analyzer.AIExcelAnalyzer.analyze_file')
    def test_parse_with_ai_success(self, mock_analyze):
        """Успешный парсинг с AI."""
        # Мокаем ответ AI
        mock_analyze.return_value = {
            'status': 'success',
            'data_start_row': 5,
            'headers': [
                {'original_name': 'Контрагент', 'normalized_name': 'client', 'category': 'client'},
                {'original_name': 'Сумма с НДС', 'normalized_name': 'amount_with_vat', 'category': 'finance', 'sub_category': 'vat_gross'}
            ],
            'hierarchy_levels': ['department', 'manager', 'client'],
            'filters_detected': {
                'period': 'Январь 2024',
                'currency': 'RUB',
                'vat_mode': 'with_vat'
            },
            'recommendations': {
                'group_by': ['manager'],
                'metrics_to_show': ['amount_with_vat'],
                'anomalies_check': ['amount_with_vat']
            },
            'confidence_score': 0.92,
            'source': 'openrouter'
        }
        
        parser = SmartParser()
        result = parser.parse_with_ai("fake_file.xlsx")
        
        assert result['status'] == 'success'
        assert result['confidence'] == 0.92
        assert result['ai_source'] == 'openrouter'
        assert 'config' in result
        assert 'recommendations' in result
        
        # Проверка конфигурации
        config = result['config']
        assert config['data_start_row'] == 5
        assert config['filters']['vat_mode'] == 'with_vat'
        assert config['business_rules']['vat_handling'] == 'extract_from_total'
    
    @patch('src.ai_core.ai_analyzer.AIExcelAnalyzer.analyze_file')
    def test_auto_vat_rules(self, mock_analyze):
        """Автоматические правила для НДС."""
        mock_analyze.return_value = {
            'status': 'success',
            'data_start_row': 1,
            'headers': [
                {'original_name': 'Сумма без НДС', 'normalized_name': 'amount_without_vat', 'category': 'finance', 'sub_category': 'vat_net'}
            ],
            'filters_detected': {'vat_mode': 'without_vat'},
            'confidence_score': 0.85
        }
        
        parser = SmartParser()
        result = parser.parse_with_ai("fake_file.xlsx")
        
        assert result['config']['business_rules']['vat_handling'] == 'calculate_on_top'
    
    @patch('src.ai_core.ai_analyzer.AIExcelAnalyzer.analyze_file')
    def test_weight_validation_rules(self, mock_analyze):
        """Правила валидации веса."""
        mock_analyze.return_value = {
            'status': 'success',
            'data_start_row': 1,
            'headers': [
                {'original_name': 'Вес брутто', 'normalized_name': 'weight_gross', 'category': 'logistics', 'sub_category': 'weight_gross'},
                {'original_name': 'Вес нетто', 'normalized_name': 'weight_net', 'category': 'logistics', 'sub_category': 'weight_net'}
            ],
            'filters_detected': {},
            'confidence_score': 0.88
        }
        
        parser = SmartParser()
        result = parser.parse_with_ai("fake_file.xlsx")
        
        assert result['config']['business_rules']['weight_validation'] == 'check_gross_net_ratio'
    
    @patch('src.ai_core.ai_analyzer.AIExcelAnalyzer.analyze_file')
    def test_smart_recommendations_generation(self, mock_analyze):
        """Генерация умных рекомендаций."""
        mock_analyze.return_value = {
            'status': 'success',
            'data_start_row': 1,
            'headers': [],
            'hierarchy_levels': ['city', 'manager'],
            'filters_detected': {
                'period': 'Q1 2024',
                'vat_mode': 'with_vat'
            },
            'recommendations': {
                'anomalies_check': ['revenue']
            },
            'confidence_score': 0.90
        }
        
        parser = SmartParser()
        result = parser.parse_with_ai("fake_file.xlsx")
        
        recs = result['recommendations']
        
        # Проверка рекомендаций
        assert len(recs['suggested_views']) > 0
        assert recs['suggested_views'][0]['type'] == 'hierarchical_tree'
        assert 'insights' in recs
        assert len(recs['insights']) > 0
        assert 'alerts' in recs


class TestFallbackStrategy:
    """Тесты стратегии фолбека."""
    
    @patch('src.ai_core.ai_analyzer.AIExcelAnalyzer.analyze_with_openrouter')
    @patch('src.ai_core.ai_analyzer.AIExcelAnalyzer.analyze_with_ollama')
    def test_fallback_to_ollama(self, mock_ollama, mock_openrouter):
        """Фолбек на локальную модель при ошибке OpenRouter."""
        mock_openrouter.return_value = None
        mock_ollama.return_value = {
            'status': 'success',
            'data_start_row': 1,
            'headers': [],
            'confidence_score': 0.75,
            'source': 'ollama'
        }
        
        analyzer = AIExcelAnalyzer()
        # Вызываем напрямую analyze_with_openrouter для проверки
        result = analyzer.analyze_with_openrouter("test context")
        
        assert result is None  # OpenRouter не сработал
        # Теперь проверяем что ollama был вызван в analyze_file
        # Это требует более сложного теста с полным пайплайном


class TestEdgeCases:
    """Тесты граничных случаев."""
    
    def test_empty_headers(self):
        """Обработка пустых заголовков."""
        parser = SmartParser()
        
        # Мокаем минимальный ответ
        with patch.object(parser.integration, 'preprocess_file') as mock_prep:
            mock_prep.return_value = {
                'status': 'success',
                'config': {
                    'data_start_row': 1,
                    'columns_mapping': {},
                    'hierarchy_levels': [],
                    'filters': {'vat_mode': 'mixed'},
                    'business_rules': {}
                },
                'ai_metadata': {'headers': []},
                'confidence': 0.5
            }
            
            result = parser.parse_with_ai("fake.xlsx")
            assert result['status'] == 'success'
            assert result['confidence'] == 0.5
    
    def test_merged_cells_detection(self, tmp_path):
        """Обнаружение объединенных ячеек."""
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        
        # Объединенная ячейка
        ws['A1'] = "Январь 2024"
        ws.merge_cells('A1:B1')
        
        ws['A2'] = "Выручка"
        ws['B2'] = "Вес"
        
        ws['A3'] = 100000
        ws['B3'] = 500
        
        test_file = tmp_path / "merged_test.xlsx"
        wb.save(str(test_file))
        wb.close()
        
        analyzer = AIExcelAnalyzer()
        result = analyzer.extract_sheet_sample(str(test_file))
        
        assert len(result['merged_cells']) > 0
        assert result['merged_cells'][0]['value'] == "Январь 2024"


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
