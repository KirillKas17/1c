"""
Модуль для экспорта документов в Excel с форматированием.

Поддерживает:
- Экспорт всех документов одного типа в один файл
- Применение маппинга полей
- Форматирование Excel (фильтры, автоширина, стили)
- Множественные листы
"""

import pandas as pd
import logging
from pathlib import Path
from typing import Dict, List, Any, Optional
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Класс для экспорта данных в Excel с форматированием."""
    
    def __init__(self):
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def export_documents_group(
        self, 
        doc_type: str, 
        documents: Dict[str, Dict[str, Any]], 
        mapping: Dict[str, Any],
        output_path: Path
    ) -> int:
        """
        Экспортирует все документы одного типа в один Excel файл.
        
        Args:
            doc_type: Тип документов
            documents: Словарь документов {filename: data}
            mapping: Маппинг полей {cell: field_name}
            output_path: Путь для сохранения файла
            
        Returns:
            Количество экспортированных документов
        """
        from core_parser.utils.validators import validate_filename, sanitize_filename
        
        logger.info(f"Экспорт {len(documents)} документов типа '{doc_type}' в {output_path}")
        
        # Валидация входных данных
        if not documents:
            logger.warning(f"Нет документов для экспорта типа '{doc_type}'")
            return 0
        
        if not mapping or not isinstance(mapping, dict):
            logger.error(f"Некорректный маппинг для типа '{doc_type}'")
            return 0
        
        # Определяем режим работы
        is_existing_file = mapping.get('target') == 'existing'
        sheet_name = mapping.get('sheet', 'Лист1')
        
        # Валидация имени листа (Excel ограничение 31 символ)
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]
            logger.warning(f"Имя листа обрезано до 31 символа: {sheet_name}")
        
        excel_path = output_path if is_existing_file else output_path.with_suffix('.xlsx')
        
        # Собираем данные для экспорта
        rows_data = []
        for filename, doc_data in documents.items():
            # Пропускаем документы с ошибками
            if isinstance(doc_data, dict) and doc_data.get('error'):
                logger.debug(f"Пропуск документа с ошибкой: {filename}")
                continue
            
            # Валидация имени файла
            try:
                safe_filename = sanitize_filename(filename)
            except Exception as e:
                logger.warning(f"Некорректное имя файла '{filename}', используем безопасное: {e}")
                safe_filename = sanitize_filename("unnamed_file")
            
            row = self._extract_row_data(doc_data, mapping)
            if row:
                row['_filename'] = safe_filename
                rows_data.append(row)
        
        if not rows_data:
            logger.warning(f"Нет данных для экспорта типа '{doc_type}'")
            return 0
        
        # Создаем DataFrame
        df = pd.DataFrame(rows_data)
        
        if is_existing_file and excel_path.exists():
            # Добавляем в существующий файл
            return self._append_to_existing_excel(df, excel_path, sheet_name, mapping)
        else:
            # Создаем новый файл
            return self._create_new_excel(df, excel_path, sheet_name, mapping)
    
    def _extract_row_data(self, doc_data: Dict[str, Any], mapping: Dict[str, Any]) -> Dict[str, Any]:
        """Извлекает данные строки из документа согласно маппингу."""
        row = {}
        
        # Получаем маппинг ячеек
        cell_mapping = mapping.get('mapping', {})
        
        # Извлекаем поля
        fields = doc_data.get('fields', {})
        
        # Извлекаем данные из таблиц
        tables = doc_data.get('tables', [])
        table_fields = doc_data.get('table_fields', {})
        
        # Применяем маппинг
        for cell_ref, field_name in cell_mapping.items():
            value = None
            
            # Ищем в полях
            if field_name in fields:
                field_data = fields[field_name]
                if isinstance(field_data, dict):
                    value = field_data.get('value')
            
            # Ищем в table_fields
            elif field_name in table_fields:
                value = table_fields[field_name]
            
            # Ищем в таблицах
            elif field_name.startswith('table_'):
                column_name = field_name.replace('table_', '')
                value = self._extract_from_tables(tables, column_name)
            
            # Специальное поле - имя файла
            elif field_name == '_filename':
                value = doc_data.get('_filename', '')
            
            row[cell_ref] = value
        
        return row
    
    def _extract_from_tables(self, tables: List[List[Dict]], column_name: str) -> Optional[str]:
        """Извлекает значение из таблиц по имени колонки."""
        for table in tables:
            if not table:
                continue
            for row in table[:5]:  # Берем первые 5 строк
                if isinstance(row, dict) and column_name in row:
                    value = row[column_name]
                    if value not in [None, '', 'NaN', 'nan']:
                        # Если список значений, объединяем
                        if isinstance(value, list):
                            return '; '.join(str(v) for v in value if v)
                        return str(value)
        return None
    
    def _create_new_excel(self, df: pd.DataFrame, excel_path: Path, sheet_name: str, mapping: Dict[str, Any]) -> int:
        """Создает новый Excel файл с форматированием."""
        # Определяем порядок колонок из маппинга
        cell_mapping = mapping.get('mapping', {})
        if cell_mapping:
            # Сортируем колонки по порядку ячеек (A1, A2, B1, B2...)
            sorted_columns = sorted(cell_mapping.keys(), key=lambda x: (x[1:], x[0]))
            # Оставляем только те колонки, которые есть в DataFrame
            columns_order = [col for col in sorted_columns if col in df.columns]
            if columns_order:
                df = df[columns_order]
        
        # Сохраняем через pandas
        with pd.ExcelWriter(str(excel_path), engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Применяем форматирование
        self._format_excel_file(excel_path, sheet_name)
        
        logger.info(f"Создан Excel файл: {excel_path}")
        return len(df)
    
    def _append_to_existing_excel(self, df: pd.DataFrame, excel_path: Path, sheet_name: str, mapping: Dict[str, Any]) -> int:
        """Добавляет данные в существующий Excel файл."""
        try:
            # Загружаем существующий файл
            book = load_workbook(excel_path)
            
            # Находим или создаем лист
            if sheet_name in book.sheetnames:
                ws = book[sheet_name]
                start_row = ws.max_row + 1
            else:
                ws = book.create_sheet(sheet_name)
                start_row = 1
            
            # Определяем колонки из маппинга
            cell_mapping = mapping.get('mapping', {})
            
            # Записываем данные
            for idx, row_data in df.iterrows():
                for cell_ref, value in row_data.items():
                    if pd.isna(value):
                        continue
                    # Преобразуем ссылку ячейки (A1) в координаты
                    col_letter = cell_ref[0]
                    col_num = ord(col_letter) - ord('A') + 1
                    ws.cell(row=start_row + idx, column=col_num, value=value)
            
            book.save(excel_path)
            book.close()
            
            # Применяем форматирование
            self._format_excel_file(excel_path, sheet_name)
            
            logger.info(f"Добавлено {len(df)} строк в существующий файл: {excel_path}")
            return len(df)
        except Exception as e:
            logger.error(f"Ошибка при добавлении в существующий файл: {e}", exc_info=True)
            return 0
    
    def _format_excel_file(self, excel_path: Path, sheet_name: str):
        """Применяет форматирование к Excel файлу."""
        try:
            book = load_workbook(excel_path)
            ws = book[sheet_name]
            
            # Форматируем заголовок
            if ws.max_row > 0:
                for cell in ws[1]:
                    cell.fill = self.header_fill
                    cell.font = self.header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = self.border
                
                # Автоширина колонок
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except (TypeError, AttributeError) as e:
                            logger.debug(f"Ошибка при обработке ячейки: {e}")
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Добавляем фильтры
                if ws.max_row > 1:
                    table_range = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
                    try:
                        table = Table(displayName=f"Table_{sheet_name}", ref=table_range)
                        style = TableStyleInfo(
                            name="TableStyleMedium9",
                            showFirstColumn=False,
                            showLastColumn=False,
                            showRowStripes=True,
                            showColumnStripes=False
                        )
                        table.tableStyleInfo = style
                        ws.add_table(table)
                    except Exception as e:
                        logger.debug(f"Не удалось добавить таблицу в Excel: {e}")
                
                # Применяем границы к ячейкам
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.border = self.border
            
            book.save(excel_path)
            book.close()
            
        except Exception as e:
            logger.error(f"Ошибка при форматировании Excel: {e}", exc_info=True)

