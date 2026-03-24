"""
AI-powered Excel Analyzer using Hybrid LLM Approach.
Main: OpenRouter (GPT-4/Claude/etc.)
Fallback: Ollama (Phi-3-mini) for local/offline processing.
"""

import os
import json
import logging
from typing import Dict, Any, Optional, List
import requests
from openpyxl import load_workbook
import pandas as pd

logger = logging.getLogger(__name__)

class AIExcelAnalyzer:
    def __init__(self):
        self.openrouter_api_key = os.getenv("OPENROUTER_API_KEY")
        self.openrouter_url = "https://openrouter.ai/api/v1/chat/completions"
        self.ollama_url = os.getenv("OLLAMA_HOST", "http://localhost:11434/api/generate")
        self.fallback_model = "phi3:mini"
        
        # Системный промпт для анализа структуры 1С и Excel
        self.system_prompt = """
Ты — эксперт по анализу данных из 1С, Excel и ERP-систем. Твоя задача — проанализировать структуру предоставленного файла и извлечь метаданные для последующей аналитики.

ВХОДНЫЕ ДАННЫЕ:
1. Заголовки столбцов (первые 2-3 строки, если есть иерархия).
2. Примеры данных (первые 5-10 строк реальных данных).
3. Метаданные (название листа, возможные объединенные ячейки в шапке).

ЧТО НУЖНО СДЕЛАТЬ:
1. Определить начало таблицы данных (пропустить шапку отчета, фильтры, параметры 1С).
2. Распознать иерархию: найти уровни вложенности (Отдел -> Город -> Менеджер -> Клиент).
3. Идентифицировать типы колонок:
   - Контрагенты (Юрлица, ИП, Физлица)
   - Товары/Номенклатура
   - Финансы (Сумма, НДС, Валюта) — различать 'с НДС' и 'без НДС'.
   - Логистика (Вес брутто/нетто, Объем).
   - Даты (Периоды, Группировки по времени).
4. Обнаружить объединенные ячейки в заголовках (например, "Январь" над "Выручка" и "Вес").
5. Выявить скрытые или пустые столбцы, которые можно игнорировать.
6. Определить единицы измерения и валюту.

ВЫХОДНОЙ ФОРМАТ (СТРОГО JSON):
{
  "data_start_row": <int>, // Номер строки, где начинаются реальные данные
  "headers": [
    {
      "original_name": "<string>",
      "normalized_name": "<snake_case_string>",
      "category": "<client|product|finance|logistics|date|hierarchy|other>",
      "sub_category": "<optional: vat_gross|vat_net|weight_gross|manager|city|etc>",
      "is_merged_parent": <bool>, // Является ли заголовок частью объединенной ячейки
      "merged_children": ["<child_col_1>", "<child_col_2>"] // Если это родитель
    }
  ],
  "hierarchy_levels": ["<level_1>", "<level_2>", ...], // Порядок вложенности
  "filters_detected": {
    "period": "<string>",
    "currency": "<string>",
    "vat_mode": "<with_vat|without_vat|mixed>",
    "custom_filters": ["<filter_description>"]
  },
  "recommendations": {
    "group_by": ["<col_name>"], // Рекомендуемая группировка
    "metrics_to_show": ["<col_name>"], // Какие метрики важны
    "anomalies_check": ["<col_name>"] // Где искать аномалии
  },
  "confidence_score": <float 0.0-1.0>
}

ВАЖНО:
- Если заголовок содержит "НДС", уточни, это сумма налога или цена с налогом.
- Различай "Брутто" и "Нетто".
- Если видишь имена людей рядом с названиями компаний, попробуй определить роль (Менеджер vs Клиент).
- Игнорируй технические столбцы 1С (например, пустые ссылки, служебные идентификаторы), если они не несут смысла.
"""

    def extract_sheet_sample(self, file_path: str, max_rows: int = 15) -> Dict[str, Any]:
        """Извлекает образец данных из файла для отправки в LLM."""
        try:
            wb = load_workbook(file_path, data_only=True)
            sheet = wb.active
            
            # Собираем сырые данные (первые N строк)
            raw_data = []
            merged_info = []
            
            # Обработка объединенных ячеек
            for merged_range in sheet.merged_cells.ranges:
                mr_min_r, mr_min_c = merged_range.min_row, merged_range.min_col
                mr_max_r, mr_max_c = merged_range.max_row, merged_range.max_col
                val = sheet.cell(row=mr_min_r, column=mr_min_c).value
                if val:
                    merged_info.append({
                        "top_left": f"R{mr_min_r}C{mr_min_c}",
                        "size": f"{mr_max_r-mr_min_r+1}x{mr_max_c-mr_min_c+1}",
                        "value": str(val)
                    })

            for row_idx, row in enumerate(sheet.iter_rows(max_row=max_rows, values_only=True), 1):
                # Пропускаем полностью пустые строки в начале (возможно часть шапки)
                if row_idx <= 5 and all(cell is None for cell in row):
                    continue
                raw_data.append([str(cell) if cell is not None else "" for cell in row])
            
            wb.close()
            
            return {
                "sheet_name": sheet.title,
                "merged_cells": merged_info,
                "sample_rows": raw_data[:max_rows],
                "total_rows_extracted": len(raw_data)
            }
        except Exception as e:
            logger.error(f"Error extracting sample: {e}")
            return {"error": str(e)}

    def analyze_with_openrouter(self, context_text: str) -> Optional[Dict]:
        """Запрос к OpenRouter (основная модель)."""
        if not self.openrouter_api_key:
            logger.warning("OpenRouter API key not found.")
            return None

        headers = {
            "Authorization": f"Bearer {self.openrouter_api_key}",
            "Content-Type": "application/json",
            "HTTP-Referer": "https://your-app.com", # Требуется OpenRouter
            "X-Title": "ExcelAnalyzer"
        }
        
        payload = {
            "model": "meta-llama/llama-3-70b-instruct", # Или "anthropic/claude-3-haiku", "openai/gpt-4-turbo"
            "messages": [
                {"role": "system", "content": self.system_prompt},
                {"role": "user", "content": f"Проанализируй эту структуру Excel/1С:\n\n{context_text}"}
            ],
            "temperature": 0.1, # Низкая температура для точности JSON
            "response_format": {"type": "json_object"}
        }

        try:
            response = requests.post(self.openrouter_url, json=payload, headers=headers, timeout=30)
            response.raise_for_status()
            result = response.json()
            content = result['choices'][0]['message']['content']
            return json.loads(content)
        except Exception as e:
            logger.error(f"OpenRouter request failed: {e}")
            return None

    def analyze_with_ollama(self, context_text: str) -> Optional[Dict]:
        """Запрос к локальной Ollama (Phi-3-mini) как фолбек."""
        prompt = f"{self.system_prompt}\n\nПроанализируй эту структуру:\n{context_text}"
        
        payload = {
            "model": self.fallback_model,
            "prompt": prompt,
            "stream": False,
            "format": "json",
            "options": {
                "temperature": 0.1
            }
        }

        try:
            response = requests.post(self.ollama_url, json=payload, timeout=60) # Локальная может думать дольше
            response.raise_for_status()
            result = response.json()
            content = result.get('response', '')
            # Очистка от маркдауна, если модель его добавит
            content = content.replace("```json", "").replace("```", "").strip()
            return json.loads(content)
        except Exception as e:
            logger.error(f"Ollama request failed: {e}")
            return None

    def analyze_file(self, file_path: str) -> Dict[str, Any]:
        """Основной метод анализа файла."""
        logger.info(f"Starting AI analysis for: {file_path}")
        
        # 1. Извлечение образца
        sample = self.extract_sheet_sample(file_path)
        if "error" in sample:
            return {"status": "error", "message": sample["error"]}

        # 2. Формирование текстового контекста для LLM
        context_parts = [
            f"Sheet: {sample['sheet_name']}",
            f"Merged Cells Info: {json.dumps(sample['merged_cells'], ensure_ascii=False)}",
            "Data Sample (Row by Row):"
        ]
        
        for i, row in enumerate(sample['sample_rows']):
            context_parts.append(f"Row {i+1}: {row}")
            
        context_text = "\n".join(context_parts)

        # 3. Попытка анализа через основную модель (OpenRouter)
        result = self.analyze_with_openrouter(context_text)
        
        # 4. Фолбек на локальную модель, если основная не сработала
        if not result:
            logger.info("Falling back to local Ollama model...")
            result = self.analyze_with_ollama(context_text)
        
        # 5. Если и локальная не сработала — возвращаем дефолтную структуру
        if not result:
            logger.warning("Both AI models failed. Using default heuristic parser.")
            return self._get_default_heuristic_result(sample)

        result['status'] = 'success'
        result['source'] = 'ai_hybrid'
        logger.info(f"AI Analysis completed with confidence: {result.get('confidence_score', 0)}")
        
        return result

    def _get_default_heuristic_result(self, sample: Dict) -> Dict:
        """Простая эвристика, если ИИ недоступен."""
        return {
            "status": "success",
            "source": "heuristic_fallback",
            "data_start_row": 1,
            "headers": [], # Будет заполнено стандартным парсером
            "confidence_score": 0.5,
            "message": "AI unavailable, using standard parsing rules."
        }

# Пример использования
if __name__ == "__main__":
    # Для теста нужен реальный файл
    # analyzer = AIExcelAnalyzer()
    # res = analyzer.analyze_file("path/to/complex_1c_report.xlsx")
    # print(json.dumps(res, indent=2, ensure_ascii=False))
    pass
